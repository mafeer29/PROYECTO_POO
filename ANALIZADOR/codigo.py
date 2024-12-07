import pandas as pd # para hacer un DataFrame
import openpyxl # Para utilizar el excel deirectamente 
import unicodedata
from openpyxl import load_workbook
import matplotlib.pyplot as plt # para graficar
import numpy as np # para arreglos


## Iniciamos nuestra clase para analizar la cobertura a partir de nuestra base de datos
class AnalizadorDatos:
    def __init__(self, archivo_excel):
        # Cargar el archivo excel en un DataFrame
        self.datos = pd.read_excel(archivo_excel)

    # Iniciamos un nuevo método para obtener los atributos que necesitamos
    def obtener_atributos_clave(self):
        # Extraer las columnas clave del DataFrame
        atributos_clave = self.datos[[ 'DEPARTAMENTO', "PROVINCIA", "DISTRITO", 'CENTRO_POBLADO', "UBIGEO_CCPP", 'EMPRESA_OPERADORA', '2G', '3G', '4G', '5G',
            'HASTA_1_MBPS', 'MÁS_DE_1_MBPS', 'CANT_EB_2G', 'CANT_EB_3G', 'CANT_EB_4G', 'CANT_EB_5G'
        ]]
        return atributos_clave


# Creamos una clase que herede la anterior
class ActualizarDatos(AnalizadorDatos):
    def __init__(self, archivo_excel):
        # Inicializamos la clase base
        super().__init__(archivo_excel)
        self.atributos_cuantificados = None  # Atributo para almacenar los datos cuantificados

    # Método para cuantificar la cobertura de tecnologías 2G, 3G, 4G y 5G
    def cuantificar_cobertura(self):
        # Obtenemos los atributos clave
        self.atributos_cuantificados = self.obtener_atributos_clave().copy()

        # Sumamos los atributos que presenta cada centro poblado
        # Los valores son 1 si existe cobertura, 0 si no existe
        self.atributos_cuantificados.loc[:,'CALIFICACION'] = (
            self.atributos_cuantificados['2G']*2 + 
            self.atributos_cuantificados['3G']*3 + 
            self.atributos_cuantificados['4G']*4 + 
            self.atributos_cuantificados['5G']*5 +
            self.atributos_cuantificados["HASTA_1_MBPS"]*0.5 +
            self.atributos_cuantificados["MÁS_DE_1_MBPS"]*1 +
            self.atributos_cuantificados["CANT_EB_2G"] +
            self.atributos_cuantificados["CANT_EB_3G"] +
            self.atributos_cuantificados["CANT_EB_4G"] +
            self.atributos_cuantificados["CANT_EB_5G"]
        )
        # Redondeamos los valores de la columna 'CALIFICACION' al entero más cercano
        # y los convierte a tipo entero explícitamente para asegurar consistencia en los datos.
        # Usamos aplly para que la funcion se aplique en cada elemento de la columna o fila
        self.atributos_cuantificados["CALIFICACION"] = (
        self.atributos_cuantificados["CALIFICACION"].apply(lambda x: int(round(x)))
        )
        # Hallamos la cantidad de eb total en cada CP
        self.atributos_cuantificados["CANT_EB_TOTAL"] = self.atributos_cuantificados["CANT_EB_2G"] + self.atributos_cuantificados["CANT_EB_3G"] + self.atributos_cuantificados["CANT_EB_4G"] + self.atributos_cuantificados["CANT_EB_5G"]
        # Normalizamos los nombres de los departamentos para evitar errores
        self.atributos_cuantificados['DEPARTAMENTO'] = (
        self.atributos_cuantificados['DEPARTAMENTO'].str.upper().str.strip()
        )
        self.atributos_cuantificados['PROVINCIA'] = (
        self.atributos_cuantificados['PROVINCIA'].str.upper().str.strip()
        )
        self.atributos_cuantificados['DISTRITO'] = (
        self.atributos_cuantificados['DISTRITO'].str.upper().str.strip()
        )
        self.atributos_cuantificados['CENTRO_POBLADO'] = (
        self.atributos_cuantificados['CENTRO_POBLADO'].str.upper().str.strip()
        )

        # Retornamos los datos cuantificados
        return self.atributos_cuantificados

    # Generar reporte de atributos_cuantificados
    def generar_reporte(self):
        reporte_cobertura = "COBERTURA_MOVIL_CUANTIFICADA.xlsx"
        # Método para convertir el DataFrame a un archivo Excel
        self.atributos_cuantificados.to_excel(reporte_cobertura, index=False)
        print(f"Reporte de cobertura cuantificada generado: {reporte_cobertura}")
    
    def generar_estadistica(self):
        # Obtener estadísticas descriptivas (como media, mediana, desviación estándar, valores mínimo, máximo y percentiles)
        # de la columna 'CALIFICACION' para evaluar la distribución y variabilidad de las calificaciones en los datos
        estadisticas = self.atributos_cuantificados['CALIFICACION'].describe()

        # Restablece el índice del DataFrame de estadísticas, convirtiendo el índice (nombres de estadísticos) en una columna regular
        # y generando un nuevo índice numérico para una visualización más clara y manipulable de los datos.
        df_estadisticas = estadisticas.to_frame(name='Valor').reset_index()
        df_estadisticas.columns = ['Estadística', 'Valor']  # Renombrar columnas para mayor claridad

        # Guardar en un archivo Excel
        reporte_estadistica = "ESTADISTICA_DESCRIPTIVA.xlsx"
        # Exporta el DataFrame de estadísticas a un archivo Excel sin incluir el índice como columna adicional
        df_estadisticas.to_excel(reporte_estadistica, index=False)
        print(f"Reporte de estadística generado: {reporte_estadistica}")

class GenerarGraficas:
    def __init__(self, actualizar_datos):
        self.actualizar_datos = actualizar_datos  # Instancia de ActualizarDatos

    def generar_histograma_calificacion(self):
        try:
            calificaciones = self.actualizar_datos.atributos_cuantificados['CALIFICACION']
            # Establece el tamaño de la figura para el gráfico
            plt.figure(figsize=(10, 6))

            # Genera un histograma con 16 intervalos (bins), color de fondo, borde y transparencia
            plt.hist(calificaciones, bins=16, color='skyblue', edgecolor='black', alpha=0.7)

            # Define el título y las etiquetas para los ejes X e Y
            plt.title('Distribución de Calificaciones de Cobertura')
            plt.xlabel('Calificación')
            plt.ylabel('Frecuencia')

            # Muestra la cuadrícula solo en el eje Y con una transparencia de 0.75
            plt.grid(axis='y', alpha=0.75)

            # Guarda el gráfico generado como una imagen PNG, ajustando el tamaño del gráfico a la figura
            plt.savefig('calificaciones.png', bbox_inches='tight') 

            # Muestra el gráfico en pantalla
            plt.show()

        # En caso de que no exista la columna mencionada
        except KeyError:
            print("Error: No se han cuantificado los datos. Por favor, llama a 'cuantificar_cobertura()' primero.")
        # Cualquier otro tipo de error
        except Exception as e:
            print(f"Ocurrió un error inesperado: {e}")

    def generar_grafico_por_departamento(self):
        try:
            # Agrupamos por departamento y calculamos la media de las calificaciones para cada uno
            promedio_departamentos = self.actualizar_datos.atributos_cuantificados.groupby('DEPARTAMENTO')['CALIFICACION'].mean()

            # Crear un gráfico de barras con los promedios de calificación por departamento
            plt.figure(figsize=(10, 6))  # Establece el tamaño de la imagen
            barras = promedio_departamentos.plot(kind='bar', color='skyblue')  # Color de las barras 

            # Añadir el título y etiquetas para los ejes
            plt.title('Calificación Promedio por Departamento', fontsize=14)
            plt.xlabel('Departamento', fontsize=12)
            plt.ylabel('Calificación Promedio', fontsize=12)

            # Rotar las etiquetas de los departamentos en el eje X para mejor visibilidad
            plt.xticks(rotation=90)  # Rota las etiquetas 90 grados para mejorar la estética

            # Añadir las calificaciones sobre cada barra para mostrar el valor numérico
            for i, valor in enumerate(promedio_departamentos):
                plt.text(i, valor + 0.05, f'{valor:.2f}', ha='center', va='bottom', fontsize=10, color='black')

            # Ajusta el diseño del gráfico y lo muestra
            plt.tight_layout()
            plt.savefig('calificacion_distrito.png', bbox_inches='tight')  # Guarda el gráfico generado como imagen PNG
            plt.show()  # Muestra el gráfico en pantalla

            # Imprimir mensaje de confirmación
            print("Gráfico de calificación promedio por departamento generado y mostrado.")

        except KeyError:
            print("Error: No se han cuantificado los datos. Por favor, llama a 'cuantificar_cobertura()' primero.")
        except Exception as e:
            print(f"Ocurrió un error inesperado: {e}")

    def generar_eb_por_departamento(self):
        try:
            # Calculamos el total acumulado de estaciones base por departamento
            total_eb_dpt = self.actualizar_datos.atributos_cuantificados.groupby('DEPARTAMENTO')['CANT_EB_TOTAL'].sum()

            # Crear un gráfico de barras con los totales de estaciones base por departamento
            plt.figure(figsize=(10, 6))  # Tamaño de la imagen
            total_eb_dpt.plot(kind='bar', color='skyblue')  # Color de las barras

            # Añadir títulos y etiquetas
            plt.title('Total de Estaciones Base por Departamento', fontsize=14)
            plt.xlabel('Departamento', fontsize=12)
            plt.ylabel('Total de Estaciones Base', fontsize=12)

            # Añadir las cantidades totales de EB sobre las barras
            for i, v in enumerate(total_eb_dpt):
                plt.text(i, v + 50, str(v), ha='center', fontsize=10)  # Ajusta el valor 50 para posicionar el texto

            # Rotar etiquetas de los departamentos
            plt.xticks(rotation=90)  # Rota las etiquetas de los departamentos

            # Ajustamos y mostramos el gráfico
            plt.tight_layout()
            plt.savefig('total_eb_departamento.png', bbox_inches='tight')  # Guardar la imagen
            plt.show()  # Mostrar el gráfico
            print("Gráfico de total de estaciones base por departamento generado y mostrado.")
        except KeyError:
            print("Error: No se han cuantificado los datos. Por favor, llama a 'cuantificar_cobertura()' primero.")
        except Exception as e:
            print(f"Ocurrió un error inesperado: {e}")


    def generar_grafico_pastel_eb_total(self):
        try:
            # Calculamos el total de estaciones base para cada tecnología
            total_2g = self.actualizar_datos.atributos_cuantificados['CANT_EB_2G'].sum()
            total_3g = self.actualizar_datos.atributos_cuantificados['CANT_EB_3G'].sum()
            total_4g = self.actualizar_datos.atributos_cuantificados['CANT_EB_4G'].sum()
            total_5g = self.actualizar_datos.atributos_cuantificados['CANT_EB_5G'].sum()
            
            # Lista con los totales por tecnología
            totales_eb = [total_2g, total_3g, total_4g, total_5g]
            tecnologias = ['2G', '3G', '4G', '5G']
            
            # Crear el gráfico de pastel con ajustes visuales
            plt.figure(figsize=(8, 6))
            wedges, texts, autotexts = plt.pie(totales_eb, 
                                            labels=tecnologias, 
                                            autopct='%1.1f%%', 
                                            startangle=140, 
                                            colors=['#66b3ff', '#99ff99', '#ff9999', '#ffcc99'],  # Colores similares al gráfico de barras
                                            wedgeprops={'edgecolor': 'black', 'linewidth': 1.5},  # Bordes definidos
                                            textprops={'fontsize': 12, 'color': 'black'})
            
            # Estilo del porcentaje
            for autotext in autotexts:
                autotext.set_fontsize(14)
                autotext.set_fontweight('bold')
            
            # Título del gráfico
            plt.title('Distribución de Estaciones Base por Tecnología en el País', fontsize=16, fontweight='bold', color='navy')
            
            # Asegura que el gráfico sea circular
            plt.axis('equal')  
            plt.tight_layout()
            plt.savefig('eb_distribucion.png', bbox_inches='tight')
            plt.show()
            
        except KeyError:
            print("Error: No se han cuantificado los datos. Por favor, llama a 'cuantificar_cobertura()' primero.")
        except Exception as e:
            print(f"Ocurrió un error inesperado: {e}")

class PropuestasSoluciones(ActualizarDatos):
    def __init__(self, archivo_excel, archivo_poblacion):

        # Inicializamos la clase base
        super().__init__(archivo_excel)
        # Cargamos los datos de población
        self.datos_poblacion = pd.read_excel(archivo_poblacion)

        # Verificamos si los datos de población se han cargado correctamente
        if self.datos_poblacion is None or self.datos_poblacion.empty: # El empty nos dice que si el DataFrame está vacío ser True
            raise ValueError("Error al cargar los datos de población: el DataFrame está vacío o es None.")
        
        # Seleccionamos las columnas que necesitamos
        self.datos_poblacion = self.datos_poblacion[["Departamento", "Provincia","Distrito","Centro Poblado", "Id Centro Poblado", "Población censada"]]
        
        # Renombramos las columnas en el DataFrame de población. 
        # 'inplace=True' asegura que los cambios se apliquen directamente sobre el DataFrame original, 
        # sin necesidad de crear una copia del mismo.
        self.datos_poblacion.rename(columns={
            "Departamento": "DEPARTAMENTO",
            "Provincia": "PROVINCIA",
            "Distrito": "DISTRITO",
            "Centro Poblado": "CENTRO_POBLADO",
            "Id Centro Poblado": "UBIGEO_CCPP",
            "Población censada": "POBLACION"
        }, inplace=True)

        # Asignar valor 0 si hay datos nulos en la columna "POBLACION"
        self.datos_poblacion["POBLACION"] = self.datos_poblacion["POBLACION"].fillna(0)
        self.atributos_cuantificados = pd.read_excel("COBERTURA_MOVIL_CUANTIFICADA.xlsx") # lo que hace read_excel es leer el excel y transformarlo a un DataFrame
        try:
            # Realiza una combinación entre el DataFrame 'atributos_cuantificados' y 'datos_poblacion' 
            # usando las columnas 'DEPARTAMENTO', 'PROVINCIA', 'DISTRITO' y 'CENTRO_POBLADO' como claves.
            # Solo se incluyen filas con coincidencias en ambas tablas (unión 'inner'). 
            self.atributos_cuantificados = pd.merge(
            self.atributos_cuantificados,
            self.datos_poblacion[["DEPARTAMENTO", "PROVINCIA", "DISTRITO", "CENTRO_POBLADO", "POBLACION"]],
            on=["DEPARTAMENTO", "PROVINCIA", "DISTRITO", "CENTRO_POBLADO"],  # Especificar varias columnas como lista
            how="inner"  # Esto indica que solo queremos filas donde haya coincidencias
            )
            reporte_cuantificado = 'REPORTE_CUANTIFICADO.xlsx'
            # El resultado se actualiza en un archivo Excel sin incluir el índice.
            self.atributos_cuantificados.to_excel(reporte_cuantificado , index=False)
            print(f"Reporte con la población generado: {reporte_cuantificado}")
            if self.atributos_cuantificados.empty:
                print("Error: La fusión resultó en un DataFrame vacío.")
                return
            
            # Verificar que CCPP no tiene información de población
            self.centros_sin_poblacion = self.atributos_cuantificados[self.atributos_cuantificados["POBLACION"]==0]
            self.centros_sin_poblacion.to_excel("CCPP_SIN_INFORMACION.xlsx", index=False)

            # Ahora filtramos los datos para trabajar con los que tienen poblacion
            self.atributos_cuantificados = self.atributos_cuantificados[self.atributos_cuantificados["POBLACION"] != 0]

            # Crear un tercer archivo sin las columnas especificadas y sin "CALIFICACION"
            if "CALIFICACION" in self.atributos_cuantificados.columns:
                reporte_sin_calificacion_2 = 'REPORTE_SIN_CALIFICACION_Y_COLUMNAS_EXTRA.xlsx'
                atributos_sin_calificacion_y_columnas = self.atributos_cuantificados.drop(columns=["CALIFICACION", "DEPARTAMENTO", "PROVINCIA", "CENTRO_POBLADO", "DISTRITO", "UBIGEO_CCPP", "EMPRESA_OPERADORA"])
                atributos_sin_calificacion_y_columnas.to_excel(reporte_sin_calificacion_2, index=False)
                print(f"Estadística generada sin 'CALIFICACION' y las columnas extra: {reporte_sin_calificacion_2}")
            else:
                print("Advertencia: La columna 'CALIFICACION' no existe en el DataFrame.")
                
        except KeyError as e:
            # Esto indica si hay columnas faltantes
            print(f"Error: {e}")
        except Exception as e:
            # Esto es para otro tipo de error
            print(f"Ocurrió un error inesperado: {e}")

    def capacidad_eb_cp(self):
        # Calcula la cantidad de habitantes por antena en cada centro poblado,
        # dividiendo la población entre el número total de estaciones base.
        # Los valores infinitos (por división entre cero) y los valores NaN se reemplazan por 0.
        self.atributos_cuantificados["ALCANCE_EB"] = (
            self.atributos_cuantificados['POBLACION'] / self.atributos_cuantificados['CANT_EB_TOTAL']
        ).replace([float('inf'), float('nan')], 0)

        # Aplica una función lambda (función anónima) a cada valor en la columna 'ALCANCE_EB'
        # para redondear el valor al entero más cercano y convertirlo a tipo entero.
        self.atributos_cuantificados["ALCANCE_EB"] = (
            self.atributos_cuantificados["ALCANCE_EB"].apply(lambda x: int(round(x)))
        )

        # Calcula la población cubierta estimada como el producto de la cantidad total de estaciones base (CANT_EB_TOTAL) 
        # y un valor de cobertura promedio por estación (150 personas). 
        # Reemplaza valores infinitos o NaN con 0 para evitar errores en cálculos posteriores.
        self.atributos_cuantificados["POBLACION_CUBIERTA"] = (
            self.atributos_cuantificados['CANT_EB_TOTAL'] * 150
        ).replace([float('inf'), float('nan')], 0)

        # Redondea los valores de la columna 'POBLACION_CUBIERTA' al entero más cercano y los convierte en enteros 
        # utilizando una función anónima (lambda).
        self.atributos_cuantificados["POBLACION_CUBIERTA"] = (
            self.atributos_cuantificados["POBLACION_CUBIERTA"].apply(lambda x: int(round(x)))
        )

        # Si la población cubierta es mayor o igual a la población, asignar 0 a POBLACION_NO_CUBIERTA
        # Usamos ".loc" para acceder a un df cuando cumple ciertas condiciones
        self.atributos_cuantificados.loc[
            self.atributos_cuantificados["POBLACION_CUBIERTA"] >= self.atributos_cuantificados["POBLACION"],
            "POBLACION_NO_CUBIERTA"
        ] = 0

        # Si la población cubierta es menor que la población, calcular la diferencia y reemplazar valores inf y NaN por 0
        self.atributos_cuantificados.loc[
            self.atributos_cuantificados["POBLACION_CUBIERTA"] < self.atributos_cuantificados["POBLACION"],
            "POBLACION_NO_CUBIERTA"
        ] = (
            self.atributos_cuantificados['POBLACION'] - self.atributos_cuantificados["POBLACION_CUBIERTA"]
        ).replace([float('inf'), float('nan')], 0)

        # Hallamos las eb_necesarias dividiendo la población entre 150
        self.atributos_cuantificados["EB_NECESARIAS"] = self.atributos_cuantificados["POBLACION"] / 150

        # Asignar 1 a valores menores que 1 y mayores que 0
        self.atributos_cuantificados.loc[
            (self.atributos_cuantificados["EB_NECESARIAS"] < 1) & (self.atributos_cuantificados["EB_NECESARIAS"] > 0),
            "EB_NECESARIAS"
        ] = 1

        # Redondeo de los valores y convertir a entero
        # "np.round" es una función de numpy que redondea al valor más cercano
        # "astype(int)" esto convierte a entero el valor asegurando que no sea de tipo flotante
        self.atributos_cuantificados["EB_NECESARIAS"] = np.round(self.atributos_cuantificados["EB_NECESARIAS"]).astype(int)

        # Calculamos las estaciones base faltantes restando las necesarias menos las disponibles
        # Reemplazamos valores infinitos o nulos con 0 para evitar inconsistencias en el análisis
        self.atributos_cuantificados["EB_FALTANTES"] = (
            self.atributos_cuantificados["EB_NECESARIAS"] - self.atributos_cuantificados["CANT_EB_TOTAL"]
        ).replace([float('inf'), float('nan')], 0)

        # En caso de que las eb necesarias sean menores a las eb que ya se encuentran instaladas
        self.atributos_cuantificados.loc[
            (self.atributos_cuantificados["EB_NECESARIAS"] < self.atributos_cuantificados["CANT_EB_TOTAL"]),
            "EB_FALTANTES"
        ] = 0

        # Itera sobre cada fila del DataFrame 'atributos_cuantificados' usando 'iterrows'
        # 'index' almacena el índice de la fila actual, mientras que 'row' contiene sus valores como una Serie.
        # Esto permite evaluar y manipular datos fila por fila de manera individual.
        for index, row in self.atributos_cuantificados.iterrows():
            # Caso 1: Sin datos de población
            if row["POBLACION"] == 0:
                self.atributos_cuantificados.at[index, "PROPUESTA_EB"] = (
                    f"No hay registros de población en el centro poblado de {row['CENTRO_POBLADO']}, se recomienda obtener datos sobre la población para poder realizar un mejor análisis"
                )

            # Caso 2: Población mayor a 150
            elif row["POBLACION"] > 150:
                if row["ALCANCE_EB"] <= 150 and row["ALCANCE_EB"] > 0:
                    self.atributos_cuantificados.at[index, "PROPUESTA_EB"] = (
                        f"La cantidad de EB ({row['CANT_EB_TOTAL']}) es suficiente y óptima para la población en el centro poblado de {row['DEPARTAMENTO']}."
                    )
                else:
                    self.atributos_cuantificados.at[index, "PROPUESTA_EB"] = (
                        f"El alcance de las estaciones base en el centro poblado de {row['CENTRO_POBLADO']} es superior a 150, por lo que se recomienda aumentar la cantidad de EB a {row['EB_NECESARIAS']} para mejorar la cobertura."
                    )
            
            # Caso 3: Población menor o igual a 150
            elif row["POBLACION"] <= 150:
                if row["CANT_EB_TOTAL"] == 0:
                    self.atributos_cuantificados.at[index, "PROPUESTA_EB"] = (
                        f"La población en el centro poblado de {row['CENTRO_POBLADO']} es muy baja, por lo que se recomienda instalar un repetidor para compartir cobertura con otro lugar. No obstante, si se desea evitar la dependencia entre estos puntos, sería más conveniente adquirir una estación base (EB) adicional."
                    )
                else:
                    self.atributos_cuantificados.at[index, "PROPUESTA_EB"] = (
                        f"La cantidad de EB ({row['CANT_EB_TOTAL']}) es suficiente y óptima para la población en el centro poblado de {row['CENTRO_POBLADO']}."
                    )
        
        # Eliminar las filas duplicadas antes de guardar el reporte
        self.atributos_cuantificados = self.atributos_cuantificados.drop_duplicates()

        # Imprimir para verificar
        print(self.atributos_cuantificados.head())

        # Guardar el archivo Excel
        self.atributos_cuantificados.to_excel("REPORTE_CENTRO_POBLADO.xlsx", index=False)

        # Abrir el archivo Excel previamente guardado con pandas utilizando openpyxl
        # wb representa el workbook (archivo completo de Excel).
        # ws representa el worksheet (hoja de trabajo activa) dentro del archivo Excel.
        wb = openpyxl.load_workbook("REPORTE_CENTRO_POBLADO.xlsx")
        ws = wb.active  # Seleccionar la hoja activa del archivo

        # Iterar sobre todas las columnas del archivo para ajustar el ancho
        # En cada iteración, col[0] hace referencia a la primera celda de la columna,
        # y .column_letter obtiene la letra correspondiente a la columna (por ejemplo, "A", "B", "C").
        for col in ws.columns:
            max_length = 0  # Inicializar variable para el largo máximo de la celda en la columna
            column = col[0].column_letter  # Obtener la letra de la columna (A, B, C, etc.)
            
            # Iterar sobre las celdas de la columna para encontrar la longitud máxima de contenido
            # En cada iteración, se verifica la longitud del valor de la celda. Si es mayor que la longitud
            # máxima registrada, se actualiza el valor de `max_length`. 
            for cell in col:
                # El bloque try-except se utiliza
                # para evitar errores si la celda está vacía o tiene un tipo de dato no esperado.
                try:
                    # Verificar si la longitud del contenido de la celda es mayor que el largo máximo actual
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)  # Actualizar el largo máximo
                except:
                    pass  # Ignorar celdas vacías o errores de tipo de datos
            
            # Ajustar el ancho de la columna basándose en el largo máximo encontrado
            adjusted_width = (max_length + 2)  # Se agrega un pequeño margen de 2 para mejor visibilidad
            ws.column_dimensions[column].width = adjusted_width  # Ajustar el tamaño de la columna

        # Guardar el archivo Excel con las columnas ajustadas
        wb.save("REPORTE_CENTRO_POBLADO.xlsx")
        wb.save("REPORTE.xlsx")

    def capacidad_eb_distrito(self):
        # Creamos un nuevo dataframe
        self.reporte_distrito = self.atributos_cuantificados.copy()

        # Calificación promedio por distrito
        self.reporte_distrito["CALIFICACION"] = self.atributos_cuantificados.groupby('DISTRITO')['CALIFICACION'].transform('mean')
        self.reporte_distrito["CALIFICACION"] = self.reporte_distrito["CALIFICACION"].apply(lambda x: int(round(x)))

        # Sumar la cantidad de estaciones base por tecnología
        self.reporte_distrito["CANT_EB_2G"] = self.atributos_cuantificados.groupby("DISTRITO")["CANT_EB_2G"].transform("sum")
        self.reporte_distrito["CANT_EB_3G"] = self.atributos_cuantificados.groupby("DISTRITO")["CANT_EB_3G"].transform("sum")
        self.reporte_distrito["CANT_EB_4G"] = self.atributos_cuantificados.groupby("DISTRITO")["CANT_EB_4G"].transform("sum")
        self.reporte_distrito["CANT_EB_5G"] = self.atributos_cuantificados.groupby("DISTRITO")["CANT_EB_5G"].transform("sum")
        self.reporte_distrito["CANT_EB_TOTAL"] = self.atributos_cuantificados.groupby("DISTRITO")["CANT_EB_TOTAL"].transform("sum")

        # Sumar la población total por distrito
        self.reporte_distrito["POBLACION_TOTAL"] = self.atributos_cuantificados.groupby("DISTRITO")["POBLACION"].transform("sum")

        # Calcular habitantes por antena en cada centro poblado
        self.reporte_distrito["ALCANCE_EB"] = (self.reporte_distrito['POBLACION_TOTAL'] / self.reporte_distrito['CANT_EB_TOTAL']).replace([float('inf'), float('nan')], 0)
        self.reporte_distrito["ALCANCE_EB"] = self.reporte_distrito["ALCANCE_EB"].apply(lambda x: int(round(x)))

        # Calcular población cubierta (asumimos 150 personas por cada antena base)
        self.reporte_distrito["POBLACION_CUBIERTA"] = (self.reporte_distrito['CANT_EB_TOTAL'] * 150).replace([float('inf'), float('nan')], 0)
        self.reporte_distrito["POBLACION_CUBIERTA"] = self.reporte_distrito["POBLACION_CUBIERTA"].apply(lambda x: int(round(x)))

        # Si la población cubierta es mayor o igual que la población total, asignar 0 a población no cubierta
        self.reporte_distrito.loc[self.reporte_distrito["POBLACION_CUBIERTA"] >= self.reporte_distrito["POBLACION_TOTAL"], "POBLACION_NO_CUBIERTA"] = 0

        # Si la población cubierta es menor que la población total, calcular la diferencia
        self.reporte_distrito.loc[self.reporte_distrito["POBLACION_CUBIERTA"] < self.reporte_distrito["POBLACION_TOTAL"], "POBLACION_NO_CUBIERTA"] = (self.reporte_distrito['POBLACION_TOTAL'] - self.reporte_distrito["POBLACION_CUBIERTA"]).replace([float('inf'), float('nan')], 0)

        # Calcular el número de estaciones base necesarias para cubrir la población total
        self.reporte_distrito["EB_NECESARIAS"] = self.reporte_distrito['POBLACION_TOTAL'] / 150

        # Asignar 1 a valores menores que 1 y mayores que 0
        self.reporte_distrito.loc[(self.reporte_distrito["EB_NECESARIAS"] < 1) & (self.reporte_distrito["EB_NECESARIAS"] > 0), "EB_NECESARIAS"] = 1

        # Redondeo de los valores y convertir a entero
        self.reporte_distrito["EB_NECESARIAS"] = np.round(self.reporte_distrito["EB_NECESARIAS"]).astype(int)

        # Calcular las estaciones base faltantes
        self.reporte_distrito["EB_FALTANTES"] = (self.reporte_distrito["EB_NECESARIAS"] - self.reporte_distrito["CANT_EB_TOTAL"]).replace([float('inf'), float('nan')], 0)

        # En caso de que las eb necesarias sean menores a las eb que ya se encuentran instaladas
        self.reporte_distrito.loc[
            (self.reporte_distrito["EB_NECESARIAS"] < self.reporte_distrito["CANT_EB_TOTAL"]),
            "EB_FALTANTES"
        ] = 0

        # Eliminar las columnas innecesarias
        columnas_a_eliminar = ['CENTRO_POBLADO', 'UBIGEO_CCPP', 'EMPRESA_OPERADORA', '2G', '3G', '4G', '5G', 'HASTA_1_MBPS', 'MÁS_DE_1_MBPS', 'POBLACION', "PROPUESTA_EB"]
        self.reporte_distrito.drop(columnas_a_eliminar, axis=1, inplace=True)

        # Eliminar filas duplicadas
        self.reporte_distrito = self.reporte_distrito.drop_duplicates()

        # Inicializar la columna "PROPUESTA_EB"
        self.reporte_distrito["PROPUESTA_EB"] = None


        # Evaluar las propuestas de estaciones base
        for index, row in self.reporte_distrito.iterrows():
            # Caso 1: Sin datos de población
            if row["POBLACION_TOTAL"] == 0:
                self.reporte_distrito.at[index, "PROPUESTA_EB"] = (
                    f"No hay registros de población en el distrito de {row['DISTRITO']}, se recomienda obtener datos sobre la población para poder realizar un mejor análisis"
                )
            
            # Caso 2: Población mayor a 150
            elif row["POBLACION_TOTAL"] > 150:
                if row["ALCANCE_EB"] <= 150 and row["ALCANCE_EB"] > 0:
                    self.reporte_distrito.at[index, "PROPUESTA_EB"] = (
                        f"La cantidad de EB ({row['CANT_EB_TOTAL']}) es suficiente y óptima para la población en el distrito de {row['DISTRITO']}."
                    )
                else:
                    self.reporte_distrito.at[index, "PROPUESTA_EB"] = (
                        f"El alcance de las estaciones base en el distrito de {row['DISTRITO']} es superior a 150, por lo que se recomienda aumentar la cantidad de EB a {row['EB_NECESARIAS']} para mejorar la cobertura."
                    )
            
            # Caso 3: Población menor o igual a 150
            elif row["POBLACION_TOTAL"] <= 150:
                if row["CANT_EB_TOTAL"] == 0:
                    self.reporte_distrito.at[index, "PROPUESTA_EB"] = (
                        f"La población en el distrito de {row['DISTRITO']} es muy baja, por lo que se recomienda instalar un repetidor para compartir cobertura con otro lugar. No obstante, si se desea evitar la dependencia entre estos puntos, sería más conveniente adquirir una estación base (EB) adicional."
                    )
                else:
                    self.reporte_distrito.at[index, "PROPUESTA_EB"] = (
                        f"La cantidad de EB ({row['CANT_EB_TOTAL']}) es suficiente y óptima para la población en el distrito de {row['DISTRITO']}."
                    )

        # Eliminar filas duplicadas
        self.reporte_distrito = self.reporte_distrito.drop_duplicates()

        # Imprimir para verificar
        print(self.reporte_distrito.head())

        # Guardar el reporte a Excel
        self.reporte_distrito.to_excel("REPORTE_DISTRITO.xlsx", index=False)

        # Ajustar el tamaño de las columnas del Excel
        wb = openpyxl.load_workbook("REPORTE_DISTRITO.xlsx")
        ws = wb.active
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Guardar el archivo Excel con las columnas ajustadas
        wb.save("REPORTE_DISTRITO.xlsx")

    def capacidad_eb_provincia(self):
        # Creamos un nuevo dataframe
        self.reporte_provincia = self.atributos_cuantificados.copy()
        self.reporte_provincia["CALIFICACION"] = self.atributos_cuantificados.groupby('PROVINCIA')['CALIFICACION'].transform('mean')
        self.reporte_provincia["CALIFICACION"] = (
            self.reporte_provincia["CALIFICACION"].apply(lambda x: int(round(x)))
        )
        self.reporte_provincia["CANT_EB_2G"] = self.atributos_cuantificados.groupby("PROVINCIA")["CANT_EB_2G"].transform("sum")
        self.reporte_provincia["CANT_EB_3G"] = self.atributos_cuantificados.groupby("PROVINCIA")["CANT_EB_3G"].transform("sum")
        self.reporte_provincia["CANT_EB_4G"] = self.atributos_cuantificados.groupby("PROVINCIA")["CANT_EB_4G"].transform("sum")
        self.reporte_provincia["CANT_EB_5G"] = self.atributos_cuantificados.groupby("PROVINCIA")["CANT_EB_5G"].transform("sum")
        self.reporte_provincia["CANT_EB_TOTAL"] = self.atributos_cuantificados.groupby("PROVINCIA")["CANT_EB_TOTAL"].transform("sum")
        self.reporte_provincia["POBLACION_TOTAL"] = self.atributos_cuantificados.groupby("PROVINCIA")["POBLACION"].transform("sum")
        
        # Calcular habitantes por antena en cada centro poblado
        self.reporte_provincia["ALCANCE_EB"] = (
            self.reporte_provincia['POBLACION_TOTAL'] /
            self.reporte_provincia['CANT_EB_TOTAL']
        ).replace([float('inf'), float('nan')], 0)

        self.reporte_provincia["ALCANCE_EB"] = (
            self.reporte_provincia["ALCANCE_EB"].apply(lambda x: int(round(x)))
        )

        self.reporte_provincia["POBLACION_CUBIERTA"] = (
            self.reporte_provincia['CANT_EB_TOTAL'] * 150
        ).replace([float('inf'), float('nan')], 0)

        self.reporte_provincia["POBLACION_CUBIERTA"] = (
            self.reporte_provincia["POBLACION_CUBIERTA"].apply(lambda x: int(round(x)))
        )
        
        self.reporte_provincia.loc[
            self.reporte_provincia["POBLACION_CUBIERTA"] >= self.reporte_provincia["POBLACION_TOTAL"],
            "POBLACION_NO_CUBIERTA"
        ] = 0

        self.reporte_provincia.loc[
            self.reporte_provincia["POBLACION_CUBIERTA"] < self.reporte_provincia["POBLACION_TOTAL"],
            "POBLACION_NO_CUBIERTA"
        ] = (
            self.reporte_provincia['POBLACION_TOTAL'] - self.reporte_provincia["POBLACION_CUBIERTA"]
        ).replace([float('inf'), float('nan')], 0)
        
        self.reporte_provincia["EB_NECESARIAS"] = self.reporte_provincia['POBLACION_TOTAL'] / 150

        # Asignar 1 a valores menores que 1 y mayores que 0
        self.reporte_provincia.loc[
            (self.reporte_provincia["EB_NECESARIAS"] < 1) & (self.reporte_provincia["EB_NECESARIAS"] > 0),
            "EB_NECESARIAS"
        ] = 1

        self.reporte_provincia["EB_NECESARIAS"] = np.round(self.reporte_provincia["EB_NECESARIAS"]).astype(int)
        self.reporte_provincia["EB_FALTANTES"] = (
            self.reporte_provincia["EB_NECESARIAS"] - self.reporte_provincia["CANT_EB_TOTAL"]
        ).replace([float('inf'), float('nan')], 0)

        # En caso de que las eb necesarias sean menores a las eb que ya se encuentran instaladas
        self.reporte_provincia.loc[
            (self.reporte_provincia["EB_NECESARIAS"] < self.reporte_provincia["CANT_EB_TOTAL"]),
            "EB_FALTANTES"
        ] = 0
        
        # Eliminar columnas innecesarias
        columnas_a_eliminar = [
            'DISTRITO', 'CENTRO_POBLADO', 'UBIGEO_CCPP', 
            'EMPRESA_OPERADORA', '2G', '3G', '4G', '5G', 'HASTA_1_MBPS', 'MÁS_DE_1_MBPS', 'POBLACION', "PROPUESTA_EB"
        ]
        self.reporte_provincia.drop(columnas_a_eliminar, axis=1, inplace=True)
        
        # Eliminar filas duplicadas
        self.reporte_provincia = self.reporte_provincia.drop_duplicates()
        self.reporte_provincia["PROPUESTA_EB"] = None
        for index, row in self.reporte_provincia.iterrows():
            # Caso 1: Sin datos de población
            if row["POBLACION_TOTAL"] == 0:
                self.reporte_provincia.at[index, "PROPUESTA_EB"] = (
                    f"No hay registros de población en la provincia de {row['PROVINCIA']}, se recomienda obtener datos sobre la población para poder realizar un mejor análisis"
                )
            
            # Caso 2: Población mayor a 150
            elif row["POBLACION_TOTAL"] > 150:
                if row["ALCANCE_EB"] <= 150 and row["ALCANCE_EB"] > 0:
                    self.reporte_provincia.at[index, "PROPUESTA_EB"] = (
                        f"La cantidad de EB ({row['CANT_EB_TOTAL']}) es suficiente y óptima para la población en la provincia de {row['PROVINCIA']}."
                    )
                else:
                    self.reporte_provincia.at[index, "PROPUESTA_EB"] = (
                        f"El alcance de las estaciones base en la provincia de {row['PROVINCIA']} es superior a 150, por lo que se recomienda aumentar la cantidad de EB a {row['EB_NECESARIAS']} para mejorar la cobertura."
                    )
            
            # Caso 3: Población menor o igual a 150
            elif row["POBLACION_TOTAL"] <= 150:
                if row["CANT_EB_TOTAL"] == 0:
                    self.reporte_provincia.at[index, "PROPUESTA_EB"] = (
                        f"La población en la provincia de {row['PROVINCIA']} es muy baja, por lo que se recomienda instalar un repetidor para compartir cobertura con otro lugar. No obstante, si se desea evitar la dependencia entre estos puntos, sería más conveniente adquirir una estación base (EB) adicional."
                    )
                else:
                    self.reporte_provincia.at[index, "PROPUESTA_EB"] = (
                        f"La cantidad de EB ({row['CANT_EB_TOTAL']}) es suficiente y óptima para la población en la provincia de {row['PROVINCIA']}."
                    )
        
        # Guardar el archivo Excel con el reporte
        self.reporte_provincia.to_excel("REPORTE_PROVINCIA.xlsx", index=False)
        self.reporte_provincia.drop_duplicates()

        # Ajustar el ancho de las columnas
        self.reporte_provincia.to_excel("REPORTE_PROVINCIA.xlsx", index=False)
        wb = load_workbook("REPORTE_PROVINCIA.xlsx")
        ws = wb.active
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
        wb.save("REPORTE_PROVINCIA.xlsx")


    def capacidad_eb_departamento(self):
        # Creamos un nuevo dataframe
        self.reporte_departamento = self.atributos_cuantificados.copy()
        self.reporte_departamento["CALIFICACION"] = self.atributos_cuantificados.groupby('DEPARTAMENTO')['CALIFICACION'].transform('mean')
        self.reporte_departamento["CALIFICACION"] = (
            self.reporte_departamento["CALIFICACION"].apply(lambda x: int(round(x)))
        )
        
        # Agregar sumas de estaciones base por departamento
        self.reporte_departamento["CANT_EB_2G"] = self.atributos_cuantificados.groupby("DEPARTAMENTO")["CANT_EB_2G"].transform("sum")
        self.reporte_departamento["CANT_EB_3G"] = self.atributos_cuantificados.groupby("DEPARTAMENTO")["CANT_EB_3G"].transform("sum")
        self.reporte_departamento["CANT_EB_4G"] = self.atributos_cuantificados.groupby("DEPARTAMENTO")["CANT_EB_4G"].transform("sum")
        self.reporte_departamento["CANT_EB_5G"] = self.atributos_cuantificados.groupby("DEPARTAMENTO")["CANT_EB_5G"].transform("sum")
        self.reporte_departamento["CANT_EB_TOTAL"] = self.atributos_cuantificados.groupby("DEPARTAMENTO")["CANT_EB_TOTAL"].transform("sum")
        self.reporte_departamento["POBLACION_TOTAL"] = self.atributos_cuantificados.groupby("DEPARTAMENTO")["POBLACION"].transform("sum")
        
        # Calcular habitantes por antena en cada centro poblado
        self.reporte_departamento["ALCANCE_EB"] = (
            self.reporte_departamento["POBLACION_TOTAL"] / self.reporte_departamento["CANT_EB_TOTAL"]
        ).replace([float('inf'), float('nan')], 0)

        self.reporte_departamento["ALCANCE_EB"] = (
            self.reporte_departamento["ALCANCE_EB"].apply(lambda x: int(round(x)))
        )

        # Calcular población cubierta
        self.reporte_departamento["POBLACION_CUBIERTA"] = (
            self.reporte_departamento['CANT_EB_TOTAL'] * 150
        ).replace([float('inf'), float('nan')], 0)

        self.reporte_departamento["POBLACION_CUBIERTA"] = (
            self.reporte_departamento["POBLACION_CUBIERTA"].apply(lambda x: int(round(x)))
        )
        
        # Calcular población no cubierta
        self.reporte_departamento.loc[
            self.reporte_departamento["POBLACION_CUBIERTA"] >= self.reporte_departamento["POBLACION_TOTAL"],
            "POBLACION_NO_CUBIERTA"
        ] = 0

        self.reporte_departamento.loc[
            self.reporte_departamento["POBLACION_CUBIERTA"] < self.reporte_departamento["POBLACION_TOTAL"],
            "POBLACION_NO_CUBIERTA"
        ] = (
            self.reporte_departamento['POBLACION_TOTAL'] - self.reporte_departamento["POBLACION_CUBIERTA"]
        ).replace([float('inf'), float('nan')], 0)
        
        self.reporte_departamento["EB_NECESARIAS"] = self.reporte_departamento['POBLACION_TOTAL'] / 150
        self.reporte_departamento.loc[
            (self.reporte_departamento["EB_NECESARIAS"] < 1) & (self.reporte_departamento["EB_NECESARIAS"] > 0),
            "EB_NECESARIAS"
        ] = 1

        # Redondeo de los valores y convertir a entero
        self.reporte_departamento["EB_NECESARIAS"] = np.round(self.reporte_departamento["EB_NECESARIAS"]).astype(int)

        # Calcular faltantes de EB
        self.reporte_departamento["EB_FALTANTES"] = (
            self.reporte_departamento["EB_NECESARIAS"] - self.reporte_departamento["CANT_EB_TOTAL"]
        ).replace([float('inf'), float('nan')], 0)

        # En caso de que las eb necesarias sean menores a las eb que ya se encuentran instaladas
        self.reporte_departamento.loc[
            (self.reporte_departamento["EB_NECESARIAS"] < self.reporte_departamento["CANT_EB_TOTAL"]),
            "EB_FALTANTES"
        ] = 0
        
        # Eliminar columnas innecesarias
        columnas_a_eliminar = [
            'PROVINCIA', 'DISTRITO', 'CENTRO_POBLADO', 'UBIGEO_CCPP', 
            'EMPRESA_OPERADORA', '2G', '3G', '4G', '5G', 'HASTA_1_MBPS', 'MÁS_DE_1_MBPS', 'POBLACION', "PROPUESTA_EB"
        ]
        self.reporte_departamento.drop(columnas_a_eliminar, axis=1, inplace=True)
        
        # **Eliminar duplicados**: Eliminamos duplicados solo basados en la columna 'DEPARTAMENTO', conservando la última aparición
        self.reporte_departamento.drop_duplicates(subset=['DEPARTAMENTO'], keep='last', inplace=True)

        # Asignar propuestas para estaciones base necesarias
        for index, row in self.reporte_departamento.iterrows():
            # Caso 1: Sin datos de población
            if row["POBLACION_TOTAL"] == 0:
                self.reporte_departamento.at[index, "PROPUESTA_EB"] = (
                    f"No hay registros de población en el departamento de {row['DEPARTAMENTO']}, se recomienda obtener datos sobre la población para poder realizar un mejor análisis"
                )
            
            # Caso 2: Población mayor a 150
            elif row["POBLACION_TOTAL"] > 150:
                if row["ALCANCE_EB"] <= 150 and row["ALCANCE_EB"] > 0:
                    self.reporte_departamento.at[index, "PROPUESTA_EB"] = (
                        f"La cantidad de EB ({row['CANT_EB_TOTAL']}) es suficiente y óptima para la población en el departamento de {row['DEPARTAMENTO']}."
                    )
                else:
                    self.reporte_departamento.at[index, "PROPUESTA_EB"] = (
                        f"El alcance de las estaciones base en el departamento de {row['DEPARTAMENTO']} es superior a 150, por lo que se recomienda aumentar la cantidad de EB a {row['EB_NECESARIAS']} para mejorar la cobertura."
                    )
            
            # Caso 3: Población menor o igual a 150
            elif row["POBLACION_TOTAL"] <= 150:
                if row["CANT_EB_TOTAL"] == 0:
                    self.reporte_departamento.at[index, "PROPUESTA_EB"] = (
                        f"La población en el departamento de {row['DEPARTAMENTO']} es muy baja, por lo que se recomienda instalar un repetidor para compartir cobertura con otro lugar. No obstante, si se desea evitar la dependencia entre estos puntos, sería más conveniente adquirir una estación base (EB) adicional."
                    )
                else:
                    self.reporte_departamento.at[index, "PROPUESTA_EB"] = (
                        f"La cantidad de EB ({row['CANT_EB_TOTAL']}) es suficiente y óptima para la población en el departamento de {row['DEPARTAMENTO']}."
                    )

        # Eliminar duplicados después de asignar propuestas (en caso de que se haya agregado alguna nueva fila)
        self.reporte_departamento.drop_duplicates(subset=['DEPARTAMENTO'], keep='last', inplace=True)
        
        # Guardar el archivo Excel
        self.reporte_departamento.to_excel("REPORTE_DEPARTAMENTO.xlsx", index=False)

        # Ajustar el ancho de las columnas en el archivo Excel
        wb = openpyxl.load_workbook("REPORTE_DEPARTAMENTO.xlsx")
        ws = wb.active  # Seleccionar la hoja activa del archivo

        for col in ws.columns:
            max_length = 0  # Inicializar variable para el largo máximo de la celda en la columna
            column = col[0].column_letter  # Obtener la letra de la columna (A, B, C, etc.)
            
            # Iterar sobre las celdas de la columna para encontrar la longitud máxima de contenido
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)  # Actualizar el largo máximo
                except:
                    pass  # Ignorar celdas vacías o errores de tipo de datos
            
            # Ajustar el ancho de la columna basándose en el largo máximo encontrado
            adjusted_width = (max_length + 2)  # Se agrega un pequeño margen de 2 para mejor visibilidad
            ws.column_dimensions[column].width = adjusted_width  # Ajustar el tamaño de la columna

        # Guardar el archivo Excel con las columnas ajustadas
        wb.save("REPORTE_DEPARTAMENTO.xlsx")

# Función para quitar tildes
def quitar_tildes(texto):
    return ''.join((c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn'))
#Esta función descompone los caracteres acentuados en su forma base y los signos diacríticos

class Poblacion:
    def __init__(self, excel_path, sheet_name):
        self.excel_path = excel_path #Almacena la ruta del archivo Excel proporcionada al crear la instancia de la clase. 
        self.sheet_name = sheet_name #hoja específica dentro del archivo Excel
        self.data = None #Inicializa un atributo para almacenar todos los datos que se leerán del archivo Excel.
        self.df_total = None #Atributo para almacenar un DataFrame de la poblacion total
        self.df_reducido = None #Atributo para almacenar un DataFrame con la poblacion reducida
        self.df_reporte = None  # Agregar un atributo para el reporte

    def cargar_datos(self):
        # Cargar el archivo Excel usando una funcion de pandas
        self.data = pd.read_excel(self.excel_path, sheet_name=self.sheet_name, header=0)

        # Renombrar columnas si la primera aparece como NaN
        #Si la primera columna no tiene un nombre válido (NaN), se le asigna un nombre genérico más descriptivo.
        if pd.isna(self.data.columns[0]):
            self.data.rename(columns={self.data.columns[0]: 'Departamento, provincia, area urbana y rural; y sexo'}, inplace=True)

        # Limpiar nombres de columnas (espacios en blanco, saltos de linea, etc)
        self.data.columns = self.data.columns.str.strip().str.replace('\n', '')

        # Eliminar tildes en los nombres de las columnas
        self.data.columns = [quitar_tildes(col) for col in self.data.columns]

        # Agregar columnas para DEPARTAMENTO y PROVINCIA
        self.data['DEPARTAMENTO'] = None
        self.data['PROVINCIA'] = None

        # Extraer departamentos y provincias
        provincia_actual = None #Almacena temporalmente el nombre de la provincia detectada en el proceso de detección
        departamento_actual = None

        #Recorre fila por fila del DataFrame usando el método iterrows()
        for index, row in self.data.iterrows():
            columna_principal = row['Departamento, provincia, area urbana y rural; y sexo']
            #Se extrae el contenido de la columna para la fila actual y se almacena en columna_principal.

            #Se verifica que el valor de columna_principal sea una cadena de texto (str).
            if isinstance(columna_principal, str):
                columna_principal = columna_principal.strip() #Elimina espacios en blanco al inicio y al final del texto.

                # Identificar departamentos y provincias
                if columna_principal.isupper() and "URBANA" not in columna_principal and "RURAL" not in columna_principal:
                    if "DEPARTAMENTO" in columna_principal: #Si contiene "DEPARTAMENTO"
                        departamento_actual = columna_principal.replace("DEPARTAMENTO ", "").strip()#Extrae el nombre del departamento eliminando la palabra "DEPARTAMENTO". 
                        provincia_actual = None  # Reiniciar la provincia cuando se detecta un departamento
                    else:
                        provincia_actual = columna_principal.strip() #Si no es un departamento pero sigue en mayúsculas, se considera una provincia.
                else:
                    # Asignar provincia y departamento
                    self.data.at[index, 'PROVINCIA'] = provincia_actual #Asigna el valor actual de provincia_actual a la columna 'PROVINCIA' en la fila actual.
                    self.data.at[index, 'DEPARTAMENTO'] = departamento_actual #Asigna el valor actual de departamento_actual a la columna 'DEPARTAMENTO' en la fila actual.

        # Eliminar la palabra 'PROVINCIA' en los valores de la columna 'PROVINCIA'
        self.data['PROVINCIA'] = self.data['PROVINCIA'].str.replace('PROVINCIA ', '', regex=False)

        # Aplica la función quitar_tildes a todos los valores de texto en el DataFrame, asegurando que no contengan tildes.
        self.data = self.data.applymap(lambda x: quitar_tildes(x) if isinstance(x, str) else x)

    def procesar_total(self):
        #Selecciona la columna donde se encuentra la información sobre áreas urbanas o rurales, filta las palabras que contenags URBANA y RURAL
        df_filtered = self.data[self.data['Departamento, provincia, area urbana y rural; y sexo'].str.contains('URBANA|RURAL', na=False)]

        #Agrupa las filas por las combinaciones únicas de las columnas 'DEPARTAMENTO' y 'PROVINCIA'. Y las suma
        self.df_total = df_filtered.groupby(['DEPARTAMENTO', 'PROVINCIA'], as_index=False).sum()
        #Filtra el DataFrame resultante para quedarse solo con las columnas que interesan
        self.df_total = self.df_total[['DEPARTAMENTO', 'PROVINCIA', 'Total', '14 a 29', '30 a 44', '45 a 64', '65 y mas']]

    def ajustar_reducida(self):
        #Ajustar la columna '65 y mas' y recalcular los totales.
        df_filtered = self.data[self.data['Departamento, provincia, area urbana y rural; y sexo'].str.contains('URBANA|RURAL', na=False)].copy()
        #Busca las filas cuyo texto mencione "URBANA" o "RURAL". Los coloca en una copia independiente del subconjunto de datos

        def ajustar_poblacion(row):
            if 'URBANA' in row['Departamento, provincia, area urbana y rural; y sexo']:
                return round(row['65 y mas'] * 0.35)  # 35% para zonas urbanas
            elif 'RURAL' in row['Departamento, provincia, area urbana y rural; y sexo']:
                return round(row['65 y mas'] * 0.20)  # 20% para zonas rurales
            return row['65 y mas']
        #Con round(): Redondea el valor ajustado al entero más cercano.

        df_filtered['65 y mas'] = df_filtered.apply(ajustar_poblacion, axis=1) #Aplica la función ajustar_poblacion fila por fila (axis=1). Modificando la columna de '65 a más'
        df_filtered['Total'] = df_filtered[['14 a 29', '30 a 44', '45 a 64', '65 y mas']].sum(axis=1) #Selecciona las columnas de los rangos de edad y las suma

        self.df_reducido = df_filtered.groupby(['DEPARTAMENTO', 'PROVINCIA'], as_index=False).sum() #Agrupa los datos por combinaciones únicas de 'DEPARTAMENTO' y 'PROVINCIA' y las suma
        self.df_reducido = self.df_reducido[['DEPARTAMENTO', 'PROVINCIA', 'Total', '14 a 29', '30 a 44', '45 a 64', '65 y mas']]
        #Filtra las columnas para quedarse solo con las de interés:'DEPARTAMENTO', 'PROVINCIA', 'Total' y los rangos de edad.

    def guardar_archivos(self):
        #Guardar los datos procesados en archivos Excel.
        if self.df_total is not None: #Verifica si el atributo self.df_total tiene un valor asignado
            self.df_total.to_excel('archivo_poblacion_total.xlsx', index=False) #Exporta el DataFrame
            print("Archivo 'archivo_poblacion_total.xlsx' guardado.")
        if self.df_reducido is not None: #Verifica si el atributo self.df_total tiene un valor asignado
            self.df_reducido.to_excel('archivo_poblacion_reducida.xlsx', index=False) #Exporta el DataFrame
            print("Archivo 'archivo_poblacion_reducida.xlsx' guardado.")
        
    def ejecutar_procesamiento(self, cobertura_excel_path=None):
        #Ejecutar todo el proceso: cargar, procesar y guardar.
        print("Cargando datos...")
        self.cargar_datos()

        print("Procesando población total...")
        self.procesar_total()

        print("Ajustando población reducida...")
        self.ajustar_reducida()

        print("Guardando archivos...")
        self.guardar_archivos()

    def cargar_datos_activa(self):
        """Cargar los archivos necesarios para agregar la población activa."""
        # Cargar el archivo de población reducida
        self.df_reducido = pd.read_excel('archivo_poblacion_reducida.xlsx') #Carga el archivo como un DataFrame y lo asigna al atributo self.df_reducido.

        # Cargar el archivo de reporte por provincia
        self.df_reporte = pd.read_excel('REPORTE_PROVINCIA.xlsx')

    def agregar_poblacion_activa(self):
        #Agregar la columna 'Total' de archivo_poblacion_reducida a REPORTE_PROVINCIA

        # Realizar el merge entre ambos DataFrames usando las columnas 'DEPARTAMENTO' y 'PROVINCIA'
        df_merge = pd.merge(self.df_reporte, self.df_reducido[['DEPARTAMENTO', 'PROVINCIA', 'Total']], on=['DEPARTAMENTO', 'PROVINCIA'], how='left')
        #on=['DEPARTAMENTO', 'PROVINCIA'], esto nos sirve para especificar las columnas a utilizar para la unión.
        #how='left': Lo que significa que se mantendrán todos los registros del DataFrame self.df_reporte y se agregarán los datos de self.df_reducido

        #Se crea una nueva columna llamada POBLACION_ACTIVA a partir de la columna Total resultante de la unión.
        df_merge['POBLACION_ACTIVA'] = df_merge['Total']
        df_merge.drop(columns=['Total'], inplace=True)  # Eliminar la columna 'Total' original

        # Se asigna el DataFrame resultante de la operación de unión y transformación al atributo self.df_reporte.
        self.df_reporte = df_merge

        # Sobrescribir el archivo 'REPORTE_PROVINCIA.xlsx' con la nueva columna
        self.df_reporte.to_excel('REPORTE_PROVINCIA.xlsx', index=False) #El DataFrame self.df_reporte se guarda nuevamente en el archivo
        #Evita que se guarde el índice de los DataFrames como una columna adicional en el archivo Excel.
        print("Archivo 'REPORTE_PROVINCIA.xlsx' actualizado con la columna POBLACION_ACTIVA.")


    def agregar_propietas_eb_pea(self):
        #Agregar la columna 'PROPUESTAS_EB_PEA' y asegurar que las columnas ALCANCE_EB_PEA y EB_NECESARIAS_PEA sean enteros.
        # Calcular ALCANCE_EB_PEA y EB_NECESARIAS_PEA
        self.df_reporte['ALCANCE_EB_PEA'] = self.df_reporte['POBLACION_ACTIVA'] / self.df_reporte['CANT_EB_TOTAL']
        self.df_reporte['EB_NECESARIAS_PEA'] = self.df_reporte['POBLACION_ACTIVA'] / 150

        # Reemplazar los valores NaN o inf con 0 antes de la conversión a entero
        self.df_reporte['ALCANCE_EB_PEA'] = self.df_reporte['ALCANCE_EB_PEA'].replace([float('inf'), -float('inf')], 0).fillna(0)
        self.df_reporte['EB_NECESARIAS_PEA'] = self.df_reporte['EB_NECESARIAS_PEA'].replace([float('inf'), -float('inf')], 0).fillna(0)

        # Las columnas ALCANCE_EB_PEA y EB_NECESARIAS_PEA se convierten a tipo entero.  
        self.df_reporte['ALCANCE_EB_PEA'] = self.df_reporte['ALCANCE_EB_PEA'].astype(int)
        self.df_reporte['EB_NECESARIAS_PEA'] = self.df_reporte['EB_NECESARIAS_PEA'].astype(int)

        # Función para definir la propuesta en 'PROPUESTAS_EB_PEA'
        def propuesta(row):
            if row["ALCANCE_EB_PEA"] <= 150 and row["ALCANCE_EB_PEA"] > 0:
                return f"La cantidad de EB ({row['CANT_EB_TOTAL']}) es suficiente y óptima tomando en cuenta solo la PEA para la población en la provincia de {row['PROVINCIA']}."
                    
            else:
                return f"El alcance de las estaciones base en la provincia de {row['PROVINCIA']} es superior a 150 tomando en cuenta solo la PEA, por lo que se recomienda aumentar la cantidad de EB a {row['EB_NECESARIAS_PEA']} para mejorar la cobertura."
             
        # Aplicar la función a cada fila
        self.df_reporte['PROPUESTAS_EB_PEA'] = self.df_reporte.apply(propuesta, axis=1)

        #El archivo REPORTE_PROVINCIA.xlsx se guarda con la nueva columna PROPUESTAS_EB_PEA añadida.
        self.df_reporte.to_excel('REPORTE_PROVINCIA.xlsx', index=False)
        print("Archivo 'REPORTE_PROVINCIA.xlsx' actualizado con la columna PROPUESTAS_EB_PEA.")

        # Imprimimos para verificar
        print(self.df_reporte.head()) #muestra las primeras 5 filas del DataFrame, para saber si los calauldos estan correctos
        self.df_reporte.to_excel("REPORTE_PROVINCIA.xlsx", index=False)

        # Abrir el archivo Excel previamente guardado con pandas utilizando openpyxl
        wb = openpyxl.load_workbook("REPORTE_PROVINCIA.xlsx")
        ws = wb.active  # Seleccionar la hoja activa del archivo

        # Iterar sobre todas las columnas del archivo
        for col in ws.columns:
            max_length = 0  # Inicializar variable para el largo máximo de la celda en la columna
            column = col[0].column_letter  # Obtener la letra de la columna (A, B, C, etc.)
            
            # Iterar sobre las celdas de la columna para encontrar la longitud máxima de contenido
            for cell in col:
                try:
                    # Verificar si la longitud del contenido de la celda es mayor que el largo máximo actual
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)  # Actualizar el largo máximo
                except:
                    pass  # Ignorar celdas vacías o errores de tipo de datos
            
            # Ajustar el ancho de la columna basándose en el largo máximo encontrado
            adjusted_width = (max_length + 2)  # Se agrega un pequeño margen de 2 para mejor visibilidad
            ws.column_dimensions[column].width = adjusted_width  # Ajustar el tamaño de la columna

        # Guardar el archivo Excel con las columnas ajustadas
        wb.save("REPORTE_PROVINCIA.xlsx")
        

# Ejemplo de uso:
# Primero, se debe proporcionar el nombre del archivo Excel que contiene los datos de cobertura
analizador = ActualizarDatos("COBERTURA MOVIL.xlsx")

# Cuantificamos la cobertura
atributos_cuantificados = analizador.cuantificar_cobertura()

# Generamos el reporte
analizador.generar_reporte()
analizador.generar_estadistica()

# Generamos las graficas
graficador = GenerarGraficas(analizador)
graficador.generar_histograma_calificacion()
graficador.generar_grafico_por_departamento()
graficador.generar_eb_por_departamento()
graficador.generar_grafico_pastel_eb_total()

# Proporcionamos el archivo necesario para las soluciones
solucionador = PropuestasSoluciones(
    "COBERTURA MOVIL.xlsx",
    "CCPP_INEI.xlsx"
)
solucionador.capacidad_eb_cp()
solucionador.capacidad_eb_departamento()
solucionador.capacidad_eb_distrito()
solucionador.capacidad_eb_provincia()

# Ruta de los archivos Excel
excel_path = 'Cuadros Estadístico del Tomo II.xlsx'  
sheet_name = 'PET1'  

# Crear una instancia de la clase Poblacion
poblacion = Poblacion(excel_path, sheet_name)

# Ejecutar el procesamiento
poblacion.ejecutar_procesamiento()

# Cargar los datos de población activa y agregar la columna
print("Cargando y agregando población activa...")
poblacion.cargar_datos_activa()
poblacion.agregar_poblacion_activa()

# Agregar la columna de propuestas
poblacion.agregar_propietas_eb_pea()



