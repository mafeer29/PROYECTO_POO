import pytest
import pandas as pd
import openpyxl
import os
from openpyxl import load_workbook
from FINAL import ActualizarDatos, GenerarGraficas, PropuestasSoluciones, PrediccionSituacionFutura, Poblacion

class TestCoberturaMovil:
    def __init__(self):
        self.archivo_excel = "COBERTURA MOVIL.xlsx"
        self.archivo_poblacion = "CCPP_INEI.xlsx"
    
    @pytest.fixture
    def setup_data(self):
        self.actualizar_datos = ActualizarDatos(self.archivo_excel)
        self.propuestas_soluciones = PropuestasSoluciones(self.archivo_excel, self.archivo_poblacion)
        return self.actualizar_datos, self.propuestas_soluciones

    def test_cuantificar_cobertura(self, setup_data):
        actualizar_datos, _ = setup_data
        atributos_cuantificados = actualizar_datos.cuantificar_cobertura()
        assert 'CALIFICACION' in atributos_cuantificados.columns
        assert not atributos_cuantificados.empty

    def test_generar_reporte(self, setup_data):
        actualizar_datos, _ = setup_data
        actualizar_datos.cuantificar_cobertura()  # Asegúrate de cuantificar primero
        actualizar_datos.generar_reporte()
        assert pd.read_excel("COBERTURA_MOVIL_CUANTIFICADA.xlsx").shape[0] > 0

    def test_generar_graficas(self, setup_data):
        actualizar_datos, _ = setup_data
        actualizar_datos.cuantificar_cobertura()  # Asegúrate de cuantificar primero
        graficador = GenerarGraficas(actualizar_datos)
        try:
            graficador.generar_histograma_calificacion() 
        except Exception as e:
            pytest.fail(f"Generación de histograma falló: {e}")
        try:
            graficador.generar_grafico_por_departamento() 
        except Exception as e:
            pytest.fail(f"Generación de gráfico por departamento falló: {e}")
        try:
            graficador.generar_eb_por_departamento()  
        except Exception as e:
            pytest.fail(f"Generación de histograma falló: {e}")
        try:
            graficador.generar_grafico_pastel_eb_total()  
        except Exception as e:
            pytest.fail(f"Generación de histograma falló: {e}")

    def test_propuestas_soluciones(self):
        propuesta = PropuestasSoluciones(self.archivo_excel, self.archivo_poblacion)

        # Verificar que los datos se han cargado correctamente
        assert not propuesta.atributos_cuantificados.empty, "El DataFrame de atributos cuantificados está vacío."
        assert not propuesta.datos_poblacion.empty, "El DataFrame de población está vacío."
        
        # Verificar que se generó el reporte de centros poblados
        assert "REPORTE_CENTRO_POBLADO.xlsx" in propuesta.__dict__, "El reporte de centros poblados no se ha generado."
        propuesta.capacidad_eb_cp()
        
        # Verificar que el archivo de reporte se creó
        wb = load_workbook("REPORTE_CENTRO_POBLADO.xlsx")
        assert wb.active.max_row > 1, "El reporte de centros poblados está vacío."
        
        # Verificar que las propuestas se generaron correctamente
        assert "PROPUESTA_EB" in propuesta.atributos_cuantificados.columns, "La columna de propuestas no se ha creado."
        
        # Limpiar después de la prueba (opcional)
        os.remove("REPORTE_CENTRO_POBLADO.xlsx")

    def test_procesamiento_poblacion(self):
        # Instancia de la clase Poblacion
        excel_path = 'Cuadros Estadístico del Tomo II.xlsx'
        sheet_name = 'PET1'
        poblacion = Poblacion(excel_path, sheet_name)
        
        # Archivo de cobertura móvil
        cobertura_excel_path = "COBERTURA_MOVIL_CUANTIFICADA.xlsx"
        
        # Ejecutar el procesamiento
        poblacion.ejecutar_procesamiento()
        
        # Verificar que los DataFrames no estén vacíos
        assert poblacion.df_total is not None, "df_total es None"
        assert not poblacion.df_total.empty, "df_total está vacío"
        assert poblacion.df_reducido is not None, "df_reducido es None"
        assert not poblacion.df_reducido.empty, "df_reducido está vacío"
        
        if poblacion.df_estaciones is not None:
            assert not poblacion.df_estaciones.empty, "df_estaciones está vacío"
        
        # Cargar datos activa y agregar población activa
        poblacion.cargar_datos_activa()
        poblacion.agregar_poblacion_activa()
        
        # Validar que se agregó la columna de población activa
        assert 'POBLACION_ACTIVA' in poblacion.df_reporte.columns, "No se agregó la columna 'POBLACION_ACTIVA'"
        
        # Ajustar reporte con propuestas
        poblacion.agregar_propietas_eb_pea()
        
        # Validar que las columnas necesarias se generaron correctamente
        assert 'ALCANCE_EB_PEA' in poblacion.df_reporte.columns, "No se generó 'ALCANCE_EB_PEA'"
        assert 'EB_NECESARIAS_PEA' in poblacion.df_reporte.columns, "No se generó 'EB_NECESARIAS_PEA'"
        assert 'PROPUESTAS_EB_PEA' in poblacion.df_reporte.columns, "No se generó 'PROPUESTAS_EB_PEA'"
        
        # Verificar dimensiones del Excel ajustadas
        wb = openpyxl.load_workbook("REPORTE_PROVINCIA.xlsx")
        ws = wb.active
        for col in ws.columns:
            column = col[0].column_letter
            assert ws.column_dimensions[column].width > 0, f"El ancho de la columna {column} no fue ajustado"
        
        # Verificar que los archivos se guardaron correctamente
        assert pd.read_excel('archivo_poblacion_total.xlsx').shape[0] > 0
        assert pd.read_excel('archivo_poblacion_reducida.xlsx').shape[0] > 0
        if poblacion.df_estaciones is not None:
            assert pd.read_excel('estaciones_base_por_provincia.xlsx').shape[0] > 0
        
        print("Pruebas completadas con éxito")

if __name__ == "__main__":
    pytest.main()